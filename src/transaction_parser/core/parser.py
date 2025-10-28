"""Transaction parsing and categorization logic"""

import csv
import json
import os
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class TransactionParser:
    """Parse and categorize bank/credit card transactions"""

    # Default categories - will be overridden by loaded config
    DEFAULT_EXPENSE_CATEGORIES = [
        "Rent", "Bills & Utilities", "Groceries", "Restaurants", "Laundry",
        "Gas", "Automotive", "Parking", "Personal", "Travel", "Shopping",
        "Entertainment", "Video Games", "Books", "Clothes", "Furniture",
        "Gifts", "Gym"
    ]

    DEFAULT_INCOME_CATEGORIES = ["Salary", "Bonus"]
    DEFAULT_PAYMENT_CATEGORIES = ["Card Payment", "Transfer", "Return"]
    IGNORE_CATEGORY = "Ignore"

    def __init__(self, mapping_file="category_mappings.json", log_callback=None):
        self.mapping_file = mapping_file
        self.log_callback = log_callback
        self.matched_amazon_orders = set()
        self.amazon_orders = []

        # Load config (categories + mappings)
        self._load_config()

    def _load_config(self):
        """Load categories and mappings from file"""
        if os.path.exists(self.mapping_file):
            try:
                with open(self.mapping_file, 'r') as f:
                    config = json.load(f)

                # Load categories
                self.EXPENSE_CATEGORIES = config.get('expense_categories', self.DEFAULT_EXPENSE_CATEGORIES)
                self.INCOME_CATEGORIES = config.get('income_categories', self.DEFAULT_INCOME_CATEGORIES)
                self.PAYMENT_CATEGORIES = config.get('payment_categories', self.DEFAULT_PAYMENT_CATEGORIES)

                # Load mappings
                self.mappings = config.get('mappings', {})

                self.log(f"Loaded config: {len(self.EXPENSE_CATEGORIES)} expense categories, "
                        f"{len(self.INCOME_CATEGORIES)} income categories, "
                        f"{len(self.mappings)} mappings")
            except Exception as e:
                self.log(f"Error loading config, using defaults: {e}")
                self._use_defaults()
        else:
            self._use_defaults()

    def _use_defaults(self):
        """Use default categories"""
        self.EXPENSE_CATEGORIES = self.DEFAULT_EXPENSE_CATEGORIES.copy()
        self.INCOME_CATEGORIES = self.DEFAULT_INCOME_CATEGORIES.copy()
        self.PAYMENT_CATEGORIES = self.DEFAULT_PAYMENT_CATEGORIES.copy()
        self.mappings = {}

    def save_config(self):
        """Save categories and mappings to file"""
        config = {
            'expense_categories': self.EXPENSE_CATEGORIES,
            'income_categories': self.INCOME_CATEGORIES,
            'payment_categories': self.PAYMENT_CATEGORIES,
            'mappings': self.mappings
        }

        with open(self.mapping_file, 'w') as f:
            json.dump(config, f, indent=2)

        self.log(f"Saved config to {self.mapping_file}")

    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def _load_mappings(self) -> Dict[str, str]:
        if os.path.exists(self.mapping_file):
            with open(self.mapping_file, 'r') as f:
                return json.load(f)
        return {}

    def _save_mappings(self):
        with open(self.mapping_file, 'w') as f:
            json.dump(self.mappings, f, indent=2)

    def _normalize_description(self, description: str) -> str:
        return description.lower().strip()

    def _is_amazon_transaction(self, description: str) -> bool:
        desc_upper = description.upper()
        return "AMAZON.COM" in desc_upper or "AMAZON MKTPL" in desc_upper

    def _find_amazon_order(self, date: datetime, amount: float) -> tuple:
        if not self.amazon_orders:
            return None, None

        amount_abs = abs(amount)
        matches = []

        for order in self.amazon_orders:
            if order['id'] in self.matched_amazon_orders:
                continue

            date_diff = (date.date() - order['date'].date()).days
            if 0 <= date_diff <= 7 and abs(order['total'] - amount_abs) < 0.01:
                matches.append(order)

        if not matches:
            return None, None

        order = matches[0]
        self.matched_amazon_orders.add(order['id'])
        return order['items'], order['id']

    def parse_csv_with_callback(self, file_path: str, date_col: str, desc_col: str,
                  amount_col: str, source: str, date_format: str = "%m/%d/%Y",
                  invert_amounts: bool = False, debit_col: str = None, credit_col: str = None,
                  has_header: bool = True) -> List[Dict]:
        transactions = []

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            if has_header:
                # Header-based CSV - use DictReader
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        date_str = row[date_col].strip()
                        description = row[desc_col].strip()
                        transaction = self._process_row_dict(row, date_str, description, amount_col,
                                                             debit_col, credit_col, date_format,
                                                             invert_amounts, source)
                        if transaction:
                            transactions.append(transaction)
                    except (KeyError, ValueError) as e:
                        self.log(f"Warning: Could not parse row: {e}")
                        continue
            else:
                # Headerless CSV - use regular reader with column indices
                reader = csv.reader(f)
                for row in reader:
                    try:
                        # Convert column indices from strings to integers
                        date_idx = int(date_col)
                        desc_idx = int(desc_col)
                        amount_idx = int(amount_col) if amount_col else None

                        date_str = row[date_idx].strip()
                        description = row[desc_idx].strip()
                        transaction = self._process_row_list(row, date_str, description, amount_idx,
                                                             date_format, invert_amounts, source)
                        if transaction:
                            transactions.append(transaction)
                    except (IndexError, ValueError) as e:
                        self.log(f"Warning: Could not parse row: {e}")
                        continue

        self.log("Categorizing transactions...")
        for txn in transactions:
            if txn['category'] is None:
                normalized = self._normalize_description(txn['description'])
                if normalized in self.mappings:
                    txn['category'] = self.mappings[normalized]
                else:
                    txn['category'] = "Uncategorized"

        return transactions

    def _process_row_dict(self, row: dict, date_str: str, description: str, amount_col: str,
                         debit_col: str, credit_col: str, date_format: str,
                         invert_amounts: bool, source: str) -> Optional[dict]:
        """Process a row from a header-based CSV"""
        # Handle Debit/Credit columns (e.g., Capital One format)
        if debit_col and credit_col:
            debit_str = row.get(debit_col, '').strip()
            credit_str = row.get(credit_col, '').strip()

            # Remove currency symbols and commas
            debit_str = debit_str.replace('$', '').replace(',', '')
            credit_str = credit_str.replace('$', '').replace(',', '')

            # Parse debit (expenses) and credit (payments/income)
            if debit_str:
                amount = -abs(float(debit_str))  # Debits are negative (expenses)
            elif credit_str:
                amount = abs(float(credit_str))   # Credits are positive (payments/income)
            else:
                return None  # Skip rows with no amount
        else:
            # Handle single Amount column
            amount_str = row[amount_col].strip()
            amount_str = amount_str.replace('$', '').replace(',', '').replace('"', '')

            if amount_str.startswith('(') and amount_str.endswith(')'):
                amount_str = '-' + amount_str[1:-1]

            amount = float(amount_str)

            if invert_amounts:
                amount = -amount

        date = datetime.strptime(date_str, date_format)

        is_amazon = self._is_amazon_transaction(description)
        matched_order_id = None

        if is_amazon:
            amazon_items, matched_order_id = self._find_amazon_order(date, amount)
            if amazon_items:
                description = amazon_items
                self.log(f"  Matched Amazon order: {description[:50]}...")

        return {
            'date': date,
            'description': description,
            'amount': amount,
            'category': None,
            'source': source,
            'amazon_order_id': matched_order_id
        }

    def _process_row_list(self, row: list, date_str: str, description: str, amount_idx: int,
                          date_format: str, invert_amounts: bool, source: str) -> Optional[dict]:
        """Process a row from a headerless CSV"""
        # Remove quotes and clean up data
        amount_str = row[amount_idx].strip().replace('$', '').replace(',', '').replace('"', '')

        if amount_str.startswith('(') and amount_str.endswith(')'):
            amount_str = '-' + amount_str[1:-1]

        amount = float(amount_str)

        if invert_amounts:
            amount = -amount

        # Clean up description - remove quotes
        description = description.replace('"', '')
        date_str = date_str.replace('"', '')

        date = datetime.strptime(date_str, date_format)

        is_amazon = self._is_amazon_transaction(description)
        matched_order_id = None

        if is_amazon:
            amazon_items, matched_order_id = self._find_amazon_order(date, amount)
            if amazon_items:
                description = amazon_items
                self.log(f"  Matched Amazon order: {description[:50]}...")

        return {
            'date': date,
            'description': description,
            'amount': amount,
            'category': None,
            'source': source,
            'amazon_order_id': matched_order_id
        }

    def generate_summary(self, transactions: List[Dict]) -> Tuple[List, List, List, Dict, Dict]:
        expenses = []
        income = []
        payments = []
        monthly_data = defaultdict(lambda: {'expenses': 0, 'income': 0})
        monthly_expense_breakdown = defaultdict(lambda: {cat: 0 for cat in self.EXPENSE_CATEGORIES})

        for txn in transactions:
            if txn['category'] == self.IGNORE_CATEGORY:
                continue

            month_key = txn['date'].strftime('%Y-%m')

            if txn['category'] in self.PAYMENT_CATEGORIES:
                payments.append(txn)
            elif txn['amount'] < 0:
                expenses.append(txn)
                amount = abs(txn['amount'])
                monthly_data[month_key]['expenses'] += amount
                if txn['category'] in self.EXPENSE_CATEGORIES:
                    monthly_expense_breakdown[month_key][txn['category']] += amount
            else:
                income.append(txn)
                monthly_data[month_key]['income'] += txn['amount']

        expenses.sort(key=lambda x: x['date'])
        income.sort(key=lambda x: x['date'])
        payments.sort(key=lambda x: x['date'])

        return expenses, income, payments, dict(monthly_data), dict(monthly_expense_breakdown)

    def export_to_excel(self, expenses: List[Dict], income: List[Dict],
                       payments: List[Dict], monthly_summary: Dict,
                       monthly_expense_breakdown: Dict,
                       output_file: str = "transaction_summary.xlsx"):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_summary = wb.create_sheet("Monthly Summary", 0)
        self._write_summary_sheet(ws_summary, monthly_summary)

        ws_expense_breakdown = wb.create_sheet("Monthly Expense Breakdown", 1)
        self._write_expense_breakdown_sheet(ws_expense_breakdown, monthly_expense_breakdown)

        ws_expenses = wb.create_sheet("Expenses")
        self._write_transaction_sheet(ws_expenses, expenses)

        ws_income = wb.create_sheet("Income")
        self._write_transaction_sheet(ws_income, income)

        ws_payments = wb.create_sheet("Payments")
        self._write_transaction_sheet(ws_payments, payments)

        wb.save(output_file)
        self.log(f"Excel file saved: {output_file}")

    def _write_transaction_sheet(self, ws, transactions: List[Dict]):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Date", "Description", "Amount", "Category", "Source"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, txn in enumerate(transactions, 2):
            ws.cell(row=row_idx, column=1, value=txn['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=txn['description'])
            ws.cell(row=row_idx, column=3, value=abs(txn['amount']))
            ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=4, value=txn['category'])
            ws.cell(row=row_idx, column=5, value=txn.get('source', ''))

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20

    def _write_summary_sheet(self, ws, monthly_summary: Dict):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Month", "Total Income", "Total Expenses", "Net Income"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        sorted_months = sorted(monthly_summary.keys())
        for row_idx, month in enumerate(sorted_months, 2):
            data = monthly_summary[month]
            net = data['income'] - data['expenses']

            ws.cell(row=row_idx, column=1, value=month)
            ws.cell(row=row_idx, column=2, value=data['income'])
            ws.cell(row=row_idx, column=2).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=3, value=data['expenses'])
            ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=4, value=net)
            ws.cell(row=row_idx, column=4).number_format = '$#,##0.00'

            if net < 0:
                ws.cell(row=row_idx, column=4).font = Font(color="FF0000")
            else:
                ws.cell(row=row_idx, column=4).font = Font(color="00AA00")

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_expense_breakdown_sheet(self, ws, monthly_expense_breakdown: Dict):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        cell = ws.cell(row=1, column=1, value="Month")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        for col_idx, category in enumerate(self.EXPENSE_CATEGORIES, 2):
            cell = ws.cell(row=1, column=col_idx, value=category)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        total_col = len(self.EXPENSE_CATEGORIES) + 2
        cell = ws.cell(row=1, column=total_col, value="Total")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        sorted_months = sorted(monthly_expense_breakdown.keys())
        for row_idx, month in enumerate(sorted_months, 2):
            breakdown = monthly_expense_breakdown[month]

            ws.cell(row=row_idx, column=1, value=month)

            month_total = 0
            for col_idx, category in enumerate(self.EXPENSE_CATEGORIES, 2):
                amount = breakdown.get(category, 0)
                month_total += amount
                cell = ws.cell(row=row_idx, column=col_idx, value=amount)
                cell.number_format = '$#,##0.00'

                if amount == 0:
                    cell.font = Font(color="999999")

            cell = ws.cell(row=row_idx, column=total_col, value=month_total)
            cell.number_format = '$#,##0.00'
            cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 12
        for col_idx in range(2, total_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
