"""Main GUI window for Transaction Parser"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
from datetime import datetime
import openpyxl

from ..core import TransactionParser
from ..config import get_all_bank_names, get_bank_format_by_name, detect_bank_format_from_headers


class TransactionParserGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Transaction Parser")
        self.root.geometry("800x700")
        
        self.parser = TransactionParser(log_callback=self.log_message)
        self.all_transactions = []
        self.csv_files = []
        self.pending_categorizations = []
        self.current_categorization_index = 0
        
        self.create_widgets()
    
    def create_widgets(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        import_frame = ttk.Frame(self.notebook)
        self.notebook.add(import_frame, text="Import Existing")
        self.create_import_tab(import_frame)
        
        csv_frame = ttk.Frame(self.notebook)
        self.notebook.add(csv_frame, text="Add Transactions")
        self.create_csv_tab(csv_frame)
        
        amazon_frame = ttk.Frame(self.notebook)
        self.notebook.add(amazon_frame, text="Amazon History")
        self.create_amazon_tab(amazon_frame)
        
        manage_frame = ttk.Frame(self.notebook)
        self.notebook.add(manage_frame, text="Manage Categories")
        self.create_manage_tab(manage_frame)
        
        review_frame = ttk.Frame(self.notebook)
        self.notebook.add(review_frame, text="Review Transactions")
        self.create_review_tab(review_frame)
        
        self.categorize_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.categorize_frame, text="Categorize")
        self.create_categorize_tab(self.categorize_frame)
        
        export_frame = ttk.Frame(self.notebook)
        self.notebook.add(export_frame, text="Export")
        self.create_export_tab(export_frame)
    
    def create_import_tab(self, parent):
        ttk.Label(parent, text="Import Existing Summary", font=("Arial", 16, "bold")).pack(pady=10)
        
        ttk.Label(parent, text="Load an existing transaction summary Excel file to continue adding transactions", 
                 wraplength=700).pack(pady=5)
        
        file_frame = ttk.LabelFrame(parent, text="Existing Summary File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=20)
        
        self.import_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.import_file_var, width=60).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_import_file).pack(side=tk.LEFT)
        
        ttk.Button(parent, text="Load Existing Summary", command=self.load_existing_summary,
                  style="Accent.TButton").pack(pady=20)
        
        self.import_status = ttk.Label(parent, text="No summary loaded", foreground="gray")
        self.import_status.pack(pady=5)
        
        info_frame = ttk.LabelFrame(parent, text="Instructions", padding=10)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        info_text = """Import an existing transaction summary to:
• Continue from where you left off
• Add more transaction CSVs to an existing summary
• Combine multiple months of data
• Edit and re-export updated summaries

The imported transactions will be loaded into the Review Transactions tab where you can:
• View all imported transactions
• Edit categories if needed
• Add new transaction CSVs
• Export an updated summary

Note: This loads the Expenses, Income, and Payments sheets from the Excel file."""
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT, wraplength=700).pack(pady=5)
    
    def create_manage_tab(self, parent):
        ttk.Label(parent, text="Manage Categories & Mappings", font=("Arial", 16, "bold")).pack(pady=10)
        
        # File operations
        file_frame = ttk.LabelFrame(parent, text="Configuration File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.config_file_var = tk.StringVar(value="category_mappings.json")
        ttk.Entry(file_frame, textvariable=self.config_file_var, width=45).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_config_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(file_frame, text="Load", command=self.load_config).pack(side=tk.LEFT, padx=2)
        ttk.Button(file_frame, text="Save", command=self.save_config, style="Accent.TButton").pack(side=tk.LEFT, padx=2)
        
        # Notebook for categories and mappings
        manage_notebook = ttk.Notebook(parent)
        manage_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Categories tab
        cat_frame = ttk.Frame(manage_notebook)
        manage_notebook.add(cat_frame, text="Categories")
        self.create_categories_subtab(cat_frame)
        
        # Mappings tab
        map_frame = ttk.Frame(manage_notebook)
        manage_notebook.add(map_frame, text="Mappings")
        self.create_mappings_subtab(map_frame)
    
    def create_categories_subtab(self, parent):
        # Category type selector
        type_frame = ttk.Frame(parent)
        type_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(type_frame, text="Category Type:").pack(side=tk.LEFT, padx=5)
        self.cat_type_var = tk.StringVar(value="Expense")
        ttk.Combobox(type_frame, textvariable=self.cat_type_var, 
                    values=["Expense", "Income", "Payment"], 
                    state="readonly", width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(type_frame, text="Refresh", command=self.refresh_category_list).pack(side=tk.LEFT, padx=5)
        
        # Category list
        list_frame = ttk.LabelFrame(parent, text="Categories", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        list_scroll = ttk.Scrollbar(list_frame)
        list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.category_listbox = tk.Listbox(list_frame, yscrollcommand=list_scroll.set, height=15)
        self.category_listbox.pack(fill=tk.BOTH, expand=True)
        list_scroll.config(command=self.category_listbox.yview)
        
        # Add/Delete controls
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(control_frame, text="New Category:").pack(side=tk.LEFT, padx=5)
        self.new_category_var = tk.StringVar()
        ttk.Entry(control_frame, textvariable=self.new_category_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Add", command=self.add_category, 
                  style="Accent.TButton").pack(side=tk.LEFT, padx=2)
        ttk.Button(control_frame, text="Delete Selected", command=self.delete_category).pack(side=tk.LEFT, padx=2)
        
        self.refresh_category_list()
    
    def create_mappings_subtab(self, parent):
        # Search/Filter
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=5)
        self.mapping_search_var = tk.StringVar()
        self.mapping_search_var.trace_add('write', lambda *args: self.refresh_mapping_list())
        ttk.Entry(search_frame, textvariable=self.mapping_search_var, width=40).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text="Clear", command=lambda: self.mapping_search_var.set("")).pack(side=tk.LEFT, padx=2)
        
        # Mappings tree
        tree_frame = ttk.LabelFrame(parent, text="Description → Category Mappings", padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.mapping_tree = ttk.Treeview(tree_frame, columns=("Description", "Category"), 
                                         show="headings", yscrollcommand=tree_scroll.set)
        self.mapping_tree.heading("Description", text="Description")
        self.mapping_tree.heading("Category", text="Category")
        self.mapping_tree.column("Description", width=400)
        self.mapping_tree.column("Category", width=200)
        self.mapping_tree.pack(fill=tk.BOTH, expand=True)
        tree_scroll.config(command=self.mapping_tree.yview)
        
        # Edit controls
        edit_frame = ttk.Frame(parent)
        edit_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(edit_frame, text="Delete Selected", command=self.delete_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame, text="Clear All Mappings", command=self.clear_all_mappings).pack(side=tk.LEFT, padx=5)
        
        self.mapping_count_var = tk.StringVar(value="Total mappings: 0")
        ttk.Label(edit_frame, textvariable=self.mapping_count_var).pack(side=tk.RIGHT, padx=10)
        
        self.refresh_mapping_list()
    
    def create_csv_tab(self, parent):
        ttk.Label(parent, text="Add Transaction CSV Files", font=("Arial", 16, "bold")).pack(pady=10)
        
        file_frame = ttk.LabelFrame(parent, text="CSV File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.csv_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.csv_file_var, width=60).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_csv_file).pack(side=tk.LEFT)
        
        source_frame = ttk.Frame(parent)
        source_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(source_frame, text="Source ID:").pack(side=tk.LEFT, padx=5)
        self.source_var = tk.StringVar(value="")
        ttk.Entry(source_frame, textvariable=self.source_var, width=30).pack(side=tk.LEFT, padx=5)
        
        # Bank/Card Format selector
        format_frame = ttk.LabelFrame(parent, text="Bank/Card Format", padding=10)
        format_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(format_frame, text="Select Format:").pack(side=tk.LEFT, padx=5)
        self.bank_format_var = tk.StringVar(value="")
        bank_formats = get_all_bank_names()
        self.bank_format_combo = ttk.Combobox(format_frame, textvariable=self.bank_format_var,
                                              values=bank_formats, state="readonly", width=25)
        self.bank_format_combo.set("Select a Format")  # Placeholder text
        self.bank_format_combo.pack(side=tk.LEFT, padx=5)
        self.bank_format_combo.bind("<<ComboboxSelected>>", self.on_bank_format_selected)
        ttk.Label(format_frame, text="(Auto-fills column names)",
                 font=("Arial", 8, "italic")).pack(side=tk.LEFT, padx=5)

        # Column configuration frame (hidden by default, shown only for Custom format)
        self.col_frame = ttk.LabelFrame(parent, text="Column Names", padding=10)
        # Don't pack it initially - will be shown when format is selected

        ttk.Label(self.col_frame, text="Date Column:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.date_col_var = tk.StringVar(value="Date")
        ttk.Entry(self.col_frame, textvariable=self.date_col_var, width=20).grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(self.col_frame, text="Description Column:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.desc_col_var = tk.StringVar(value="Description")
        ttk.Entry(self.col_frame, textvariable=self.desc_col_var, width=20).grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(self.col_frame, text="Amount Column:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.amount_col_var = tk.StringVar(value="Amount")
        ttk.Entry(self.col_frame, textvariable=self.amount_col_var, width=20).grid(row=2, column=1, padx=5, pady=2)

        ttk.Label(self.col_frame, text="OR use Debit/Credit columns:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)

        ttk.Label(self.col_frame, text="Debit Column:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
        self.debit_col_var = tk.StringVar(value="")
        ttk.Entry(self.col_frame, textvariable=self.debit_col_var, width=20).grid(row=4, column=1, padx=5, pady=2)

        ttk.Label(self.col_frame, text="Credit Column:").grid(row=5, column=0, sticky=tk.W, padx=5, pady=2)
        self.credit_col_var = tk.StringVar(value="")
        ttk.Entry(self.col_frame, textvariable=self.credit_col_var, width=20).grid(row=5, column=1, padx=5, pady=2)

        ttk.Label(self.col_frame, text="(Leave Debit/Credit empty to use Amount column)",
                 font=("Arial", 8, "italic")).grid(row=6, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2)

        # Date format frame (hidden by default, shown only for Custom format)
        self.date_format_frame = ttk.Frame(parent)
        # Don't pack it initially - will be shown when format is selected
        ttk.Label(self.date_format_frame, text="Date Format:").pack(side=tk.LEFT, padx=5)
        self.date_format_var = tk.StringVar(value="%m/%d/%Y")
        ttk.Combobox(self.date_format_frame, textvariable=self.date_format_var,
                    values=["%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"], width=15).pack(side=tk.LEFT, padx=5)

        # Invert amounts checkbox (hidden by default, shown only for Custom format)
        self.invert_frame = ttk.Frame(parent)
        # Don't pack it initially - will be shown when format is selected
        self.invert_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self.invert_frame, text="Inverted amounts (expenses are positive)",
                       variable=self.invert_var).pack(side=tk.LEFT)

        # All configuration fields are hidden by default
        # They will be shown when a format is selected in on_bank_format_selected()

        ttk.Button(parent, text="Add CSV File", command=self.add_csv_file,
                  style="Accent.TButton").pack(pady=10)
        
        list_frame = ttk.LabelFrame(parent, text="Added Files", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.files_listbox = tk.Listbox(list_frame, height=8)
        self.files_listbox.pack(fill=tk.BOTH, expand=True)
        
        self.csv_status = ttk.Label(parent, text="", foreground="blue")
        self.csv_status.pack(pady=5)
    
    def create_amazon_tab(self, parent):
        ttk.Label(parent, text="Amazon Order History", font=("Arial", 16, "bold")).pack(pady=10)
        
        ttk.Label(parent, text="Load your Amazon order history to automatically match purchases", 
                 wraplength=700).pack(pady=5)
        
        file_frame = ttk.LabelFrame(parent, text="Amazon Order History CSV", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=20)
        
        self.amazon_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.amazon_file_var, width=60).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_amazon_file).pack(side=tk.LEFT)
        
        col_frame = ttk.LabelFrame(parent, text="Column Names", padding=10)
        col_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(col_frame, text="Date Column:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.amazon_date_col_var = tk.StringVar(value="date")
        ttk.Entry(col_frame, textvariable=self.amazon_date_col_var, width=20).grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(col_frame, text="Total Column:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.amazon_total_col_var = tk.StringVar(value="total")
        ttk.Entry(col_frame, textvariable=self.amazon_total_col_var, width=20).grid(row=1, column=1, padx=5, pady=2)
        
        ttk.Label(col_frame, text="Items Column:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.amazon_items_col_var = tk.StringVar(value="items")
        ttk.Entry(col_frame, textvariable=self.amazon_items_col_var, width=20).grid(row=2, column=1, padx=5, pady=2)
        
        ttk.Button(parent, text="Load Amazon History", command=self.load_amazon_history,
                  style="Accent.TButton").pack(pady=20)
        
        self.amazon_status = ttk.Label(parent, text="No Amazon history loaded", foreground="gray")
        self.amazon_status.pack(pady=5)
        
        info_frame = ttk.LabelFrame(parent, text="Instructions", padding=10)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        info_text = """Amazon order history helps match generic Amazon charges to actual purchases.

Expected CSV format:
• date: Order date (MM/DD/YYYY format)
• total: Order total amount
• items: Description of items purchased

Note: Amazon history only applies to transactions added AFTER loading the history file.
To apply to existing transactions, reload your transaction CSVs after loading Amazon history."""
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT, wraplength=700).pack(pady=5)
    
    def create_review_tab(self, parent):
        ttk.Label(parent, text="Review & Edit Transactions", font=("Arial", 16, "bold")).pack(pady=10)
        
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(filter_frame, text="Filter:").pack(side=tk.LEFT, padx=5)
        self.filter_var = tk.StringVar(value="All")
        filter_combo = ttk.Combobox(filter_frame, textvariable=self.filter_var, 
                                     values=["All", "Expenses", "Income", "Payments", "Uncategorized", "Ignored"],
                                     width=15, state="readonly")
        filter_combo.pack(side=tk.LEFT, padx=5)
        filter_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_transaction_list())
        
        ttk.Button(filter_frame, text="Refresh", command=self.refresh_transaction_list).pack(side=tk.LEFT, padx=5)
        
        self.review_count_var = tk.StringVar(value="Transactions: 0")
        ttk.Label(filter_frame, textvariable=self.review_count_var).pack(side=tk.LEFT, padx=20)
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tree_scroll_y = ttk.Scrollbar(tree_frame)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.transaction_tree = ttk.Treeview(tree_frame, 
                                             columns=("Date", "Description", "Amount", "Category", "Source"),
                                             show="headings",
                                             yscrollcommand=tree_scroll_y.set,
                                             xscrollcommand=tree_scroll_x.set)
        
        tree_scroll_y.config(command=self.transaction_tree.yview)
        tree_scroll_x.config(command=self.transaction_tree.xview)
        
        self.transaction_tree.heading("Date", text="Date")
        self.transaction_tree.heading("Description", text="Description")
        self.transaction_tree.heading("Amount", text="Amount")
        self.transaction_tree.heading("Category", text="Category")
        self.transaction_tree.heading("Source", text="Source")
        
        self.transaction_tree.column("Date", width=100)
        self.transaction_tree.column("Description", width=300)
        self.transaction_tree.column("Amount", width=100)
        self.transaction_tree.column("Category", width=150)
        self.transaction_tree.column("Source", width=150)
        
        self.transaction_tree.pack(fill=tk.BOTH, expand=True)
        
        edit_frame = ttk.LabelFrame(parent, text="Edit Selected Transaction", padding=10)
        edit_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(edit_frame, text="Change Category:").pack(side=tk.LEFT, padx=5)
        
        self.edit_category_var = tk.StringVar()
        self.edit_category_combo = ttk.Combobox(edit_frame, textvariable=self.edit_category_var,
                                                values=[], width=20)
        self.edit_category_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(edit_frame, text="Update Category", command=self.update_transaction_category,
                  style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(edit_frame, text="Delete Transaction", command=self.delete_transaction).pack(side=tk.LEFT, padx=5)
    
    def create_categorize_tab(self, parent):
        ttk.Label(parent, text="Categorize Transactions", font=("Arial", 16, "bold")).pack(pady=10)
        
        info_frame = ttk.LabelFrame(parent, text="Transaction Details", padding=10)
        info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.cat_description_var = tk.StringVar(value="No transactions to categorize")
        ttk.Label(info_frame, textvariable=self.cat_description_var, wraplength=700, 
                 font=("Arial", 14)).pack(pady=5)
        
        self.cat_amount_var = tk.StringVar(value="")
        ttk.Label(info_frame, textvariable=self.cat_amount_var, font=("Arial", 14, "bold")).pack(pady=5)
        
        cat_select_frame = ttk.LabelFrame(parent, text="Select Category", padding=10)
        cat_select_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        canvas = tk.Canvas(cat_select_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(cat_select_frame, orient="vertical", command=canvas.yview)
        self.category_button_frame = ttk.Frame(canvas)
        
        self.category_button_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=self.category_button_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Make the frame expand to fill the canvas width
        def configure_canvas_window(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind('<Configure>', configure_canvas_window)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(button_frame, text="Skip This Transaction", command=self.skip_categorization).pack(side=tk.LEFT, padx=5)
        
        self.cat_progress_var = tk.StringVar(value="Progress: 0 / 0")
        ttk.Label(button_frame, textvariable=self.cat_progress_var, font=("Arial", 11)).pack(side=tk.RIGHT, padx=5)
    
    def create_export_tab(self, parent):
        ttk.Label(parent, text="Export Summary", font=("Arial", 16, "bold")).pack(pady=10)
        
        file_frame = ttk.LabelFrame(parent, text="Output File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.output_file_var = tk.StringVar(value="transaction_summary.xlsx")
        ttk.Entry(file_frame, textvariable=self.output_file_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_output_file).pack(side=tk.LEFT)
        
        ttk.Button(parent, text="Generate Excel Summary", command=self.export_summary,
                  style="Accent.TButton").pack(pady=20)
        
        log_frame = ttk.LabelFrame(parent, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, width=80)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def browse_csv_file(self):
        filename = filedialog.askopenfilename(title="Select Transaction CSV",
                                             filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if filename:
            self.csv_file_var.set(filename)

            # Auto-detect bank format from CSV headers
            try:
                with open(filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # Read first row (headers)

                    detected_format = detect_bank_format_from_headers(headers)

                    if detected_format:
                        # Auto-select the detected format
                        self.bank_format_var.set(detected_format.name)
                        # Trigger the selection handler to auto-fill fields
                        self.on_bank_format_selected()
                        self.log_message(f"Auto-detected format: {detected_format.name}")
                    else:
                        # No match found, suggest Custom
                        if self.bank_format_var.get() in ["", "Select a Format"]:
                            self.log_message("Could not auto-detect bank format. Please select a format manually.")
            except Exception as e:
                self.log_message(f"Warning: Could not read CSV headers: {e}")
    
    def browse_amazon_file(self):
        filename = filedialog.askopenfilename(title="Select Amazon Order History CSV",
                                             filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if filename:
            self.amazon_file_var.set(filename)
    
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(title="Save Excel File As",
                                               defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filename:
            self.output_file_var.set(filename)
    
    def browse_import_file(self):
        filename = filedialog.askopenfilename(title="Select Existing Transaction Summary",
                                             filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filename:
            self.import_file_var.set(filename)
    
    def browse_config_file(self):
        filename = filedialog.askopenfilename(title="Select Configuration File",
                                             filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                                             initialfile="category_mappings.json")
        if filename:
            self.config_file_var.set(filename)

    def on_bank_format_selected(self, _event=None):
        """Handle bank format selection and auto-fill column names"""
        format_name = self.bank_format_var.get()

        # Ignore if placeholder text is still shown
        if format_name == "Select a Format" or not format_name:
            return

        bank_format = get_bank_format_by_name(format_name)

        if bank_format:
            # Auto-fill column names
            self.date_col_var.set(bank_format.date_col)
            self.desc_col_var.set(bank_format.desc_col)

            if bank_format.amount_col:
                self.amount_col_var.set(bank_format.amount_col)
                self.debit_col_var.set("")
                self.credit_col_var.set("")
            else:
                self.amount_col_var.set("Amount")
                self.debit_col_var.set(bank_format.debit_col or "")
                self.credit_col_var.set(bank_format.credit_col or "")

            # Set date format
            self.date_format_var.set(bank_format.date_format)

            # Set invert amounts
            self.invert_var.set(bank_format.invert_amounts)

            # Hide/show configuration fields based on format selection
            if format_name == "Custom":
                # Show all configuration fields for custom format
                # Pack them in the correct order (after the format selector)
                self.col_frame.pack(after=self.bank_format_combo.master, fill=tk.X, padx=10, pady=5)
                self.date_format_frame.pack(after=self.col_frame, fill=tk.X, padx=10, pady=5)
                self.invert_frame.pack(after=self.date_format_frame, fill=tk.X, padx=10, pady=5)
            else:
                # Hide configuration fields for pre-configured formats
                self.col_frame.pack_forget()
                self.date_format_frame.pack_forget()
                self.invert_frame.pack_forget()

            self.log_message(f"Loaded format: {format_name}")

    def log_message(self, message):
        if hasattr(self, 'log_text'):
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.root.update()
    
    def load_existing_summary(self):
        import_file = self.import_file_var.get()
        if not import_file or not os.path.exists(import_file):
            messagebox.showerror("Error", "Please select a valid Excel file")
            return
        
        try:
            wb = openpyxl.load_workbook(import_file)
            loaded_transactions = []
            
            for sheet_name in ["Expenses", "Income", "Payments"]:
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    
                    try:
                        date_str = row[0]
                        description = row[1]
                        amount = float(row[2])
                        category = row[3]
                        source = row[4] if len(row) > 4 else "Unknown"
                        
                        if isinstance(date_str, str):
                            date = datetime.strptime(date_str, '%Y-%m-%d')
                        else:
                            date = date_str
                        
                        if sheet_name == "Expenses":
                            amount = -abs(amount)
                        else:
                            amount = abs(amount)
                        
                        loaded_transactions.append({
                            'date': date,
                            'description': description,
                            'amount': amount,
                            'category': category,
                            'source': source
                        })
                    except (ValueError, TypeError, IndexError) as e:
                        self.log_message(f"Warning: Could not parse row in {sheet_name}: {e}")
                        continue
            
            wb.close()
            
            self.all_transactions.extend(loaded_transactions)
            self.refresh_transaction_list()
            
            self.import_status.config(text=f"✓ Loaded {len(loaded_transactions)} transactions", foreground="green")
            self.log_message(f"Imported {len(loaded_transactions)} transactions from {import_file}")
            
            messagebox.showinfo("Success", 
                              f"Loaded {len(loaded_transactions)} transactions from existing summary!\n\n"
                              "You can now:\n"
                              "• Add more transaction CSVs\n"
                              "• Review and edit in the Review tab\n"
                              "• Export an updated summary")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load summary: {str(e)}")
            self.import_status.config(text="✗ Error loading file", foreground="red")
            self.log_message(f"ERROR loading summary: {str(e)}")
    
    def load_amazon_history(self):
        amazon_file = self.amazon_file_var.get()
        if not amazon_file or not os.path.exists(amazon_file):
            messagebox.showerror("Error", "Please select a valid Amazon order history CSV file")
            return
        
        try:
            orders = []
            with open(amazon_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                date_col = self.amazon_date_col_var.get()
                total_col = self.amazon_total_col_var.get()
                items_col = self.amazon_items_col_var.get()
                
                for row in reader:
                    try:
                        date_str = row.get(date_col, '').strip()
                        total_str = row.get(total_col, '').strip()
                        total_str = total_str.replace('$', '').replace(',', '')
                        items = row.get(items_col, '').strip()
                        
                        if not date_str or not total_str:
                            continue
                        
                        date = datetime.strptime(date_str, '%m/%d/%Y')
                        total = abs(float(total_str))
                        
                        orders.append({
                            'date': date,
                            'total': total,
                            'items': items,
                            'raw_row': row,
                            'id': len(orders)
                        })
                    except (ValueError, KeyError) as e:
                        self.log_message(f"Warning: Could not parse Amazon order: {e}")
                        continue
            
            self.parser.amazon_orders = orders
            self.parser.matched_amazon_orders = set()
            
            self.amazon_status.config(text=f"✓ Loaded {len(orders)} Amazon orders", foreground="green")
            self.log_message(f"Loaded {len(orders)} Amazon orders from {amazon_file}")
            messagebox.showinfo("Success", f"Loaded {len(orders)} Amazon orders!\n\n"
                              "This will apply to transaction CSVs added after this point.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Amazon history: {str(e)}")
            self.amazon_status.config(text="✗ Error loading file", foreground="red")
    
    def add_csv_file(self):
        csv_file = self.csv_file_var.get()
        if not csv_file or not os.path.exists(csv_file):
            messagebox.showerror("Error", "Please select a valid CSV file")
            return
        
        source = self.source_var.get()
        if not source:
            messagebox.showerror("Error", "Please enter a source identifier")
            return
        
        self.csv_status.config(text="Processing...", foreground="blue")
        self.root.update()
        
        try:
            debit_col = self.debit_col_var.get() if self.debit_col_var.get() else None
            credit_col = self.credit_col_var.get() if self.credit_col_var.get() else None

            transactions = self.parser.parse_csv_with_callback(
                csv_file,
                self.date_col_var.get(),
                self.desc_col_var.get(),
                self.amount_col_var.get(),
                source,
                self.date_format_var.get(),
                self.invert_var.get(),
                debit_col,
                credit_col
            )
            
            self.all_transactions.extend(transactions)
            self.csv_files.append(f"{source}: {len(transactions)} transactions")
            self.files_listbox.insert(tk.END, self.csv_files[-1])
            
            uncategorized_descriptions = {}
            for txn in transactions:
                if txn['category'] == "Uncategorized":
                    normalized = self.parser._normalize_description(txn['description'])
                    if normalized not in uncategorized_descriptions:
                        uncategorized_descriptions[normalized] = txn
            
            for txn in uncategorized_descriptions.values():
                self.pending_categorizations.append(txn)
            
            self.csv_status.config(text=f"✓ Added {len(transactions)} transactions from {source}", 
                                  foreground="green")
            self.log_message(f"Added {len(transactions)} transactions from {source}")
            
            self.refresh_transaction_list()
            
            if uncategorized_descriptions:
                unique_count = len(uncategorized_descriptions)
                messagebox.showinfo("Categorization Needed", 
                                   f"{unique_count} unique transaction type(s) need categorization.\n"
                                   "Go to the Categorize tab to continue.")
                self.notebook.select(4)  # Categorize is now the 5th tab (index 4)
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process CSV: {str(e)}")
            self.csv_status.config(text="✗ Error processing file", foreground="red")
    
    def on_tab_changed(self, event):
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        
        if tab_text == "Categorize" and self.pending_categorizations:
            self.show_current_categorization()
    
    def refresh_transaction_list(self):
        if not hasattr(self, 'transaction_tree'):
            return
        
        if self.parser and hasattr(self, 'edit_category_combo'):
            all_categories = (self.parser.EXPENSE_CATEGORIES + self.parser.INCOME_CATEGORIES + 
                             self.parser.PAYMENT_CATEGORIES + [self.parser.IGNORE_CATEGORY])
            self.edit_category_combo['values'] = all_categories
        
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        if not self.all_transactions:
            self.review_count_var.set("Transactions: 0")
            return
        
        filter_type = self.filter_var.get()
        filtered = []
        
        for txn in self.all_transactions:
            if filter_type == "All":
                filtered.append(txn)
            elif filter_type == "Expenses" and txn['amount'] < 0:
                filtered.append(txn)
            elif filter_type == "Income" and txn['amount'] > 0 and txn['category'] in self.parser.INCOME_CATEGORIES:
                filtered.append(txn)
            elif filter_type == "Payments" and txn['category'] in self.parser.PAYMENT_CATEGORIES:
                filtered.append(txn)
            elif filter_type == "Uncategorized" and txn['category'] == "Uncategorized":
                filtered.append(txn)
            elif filter_type == "Ignored" and txn['category'] == self.parser.IGNORE_CATEGORY:
                filtered.append(txn)
        
        filtered.sort(key=lambda x: x['date'], reverse=True)
        
        for txn in filtered:
            date_str = txn['date'].strftime('%Y-%m-%d')
            amount_str = f"${abs(txn['amount']):.2f}"
            if txn['amount'] < 0:
                amount_str = "-" + amount_str
            
            self.transaction_tree.insert("", tk.END, values=(
                date_str,
                txn['description'][:50] + "..." if len(txn['description']) > 50 else txn['description'],
                amount_str,
                txn['category'],
                txn['source']
            ), tags=(str(id(txn)),))
        
        self.review_count_var.set(f"Transactions: {len(filtered)} (of {len(self.all_transactions)} total)")
    
    def update_transaction_category(self):
        selection = self.transaction_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a transaction to edit")
            return
        
        new_category = self.edit_category_var.get()
        if not new_category:
            messagebox.showwarning("No Category", "Please select a new category")
            return
        
        item = self.transaction_tree.item(selection[0])
        values = item['values']
        date_str = values[0]
        description = values[1]
        
        for txn in self.all_transactions:
            if (txn['date'].strftime('%Y-%m-%d') == date_str and 
                txn['description'].startswith(description.replace("...", ""))):
                
                old_category = txn['category']
                txn['category'] = new_category
                
                normalized = self.parser._normalize_description(txn['description'])
                self.parser.mappings[normalized] = new_category
                self.parser._save_mappings()
                
                self.log_message(f"Updated: {txn['description'][:30]}... | {old_category} → {new_category}")
                messagebox.showinfo("Success", f"Category updated to: {new_category}")
                
                self.refresh_transaction_list()
                break
    
    def delete_transaction(self):
        selection = self.transaction_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a transaction to delete")
            return
        
        response = messagebox.askyesno("Confirm Delete", 
                                      "Are you sure you want to delete this transaction?\n"
                                      "This cannot be undone.")
        if not response:
            return
        
        item = self.transaction_tree.item(selection[0])
        values = item['values']
        date_str = values[0]
        description = values[1]
        
        for i, txn in enumerate(self.all_transactions):
            if (txn['date'].strftime('%Y-%m-%d') == date_str and 
                txn['description'].startswith(description.replace("...", ""))):
                
                deleted_txn = self.all_transactions.pop(i)
                self.log_message(f"Deleted: {deleted_txn['description'][:50]}...")
                messagebox.showinfo("Deleted", "Transaction deleted successfully")
                
                self.refresh_transaction_list()
                break
    
    def show_current_categorization(self):
        if self.current_categorization_index >= len(self.pending_categorizations):
            messagebox.showinfo("Complete", "All transactions have been categorized!")
            self.cat_description_var.set("All transactions categorized!")
            self.cat_amount_var.set("")
            for widget in self.category_button_frame.winfo_children():
                widget.destroy()
            self.pending_categorizations = []
            self.current_categorization_index = 0
            return
        
        txn = self.pending_categorizations[self.current_categorization_index]
        self.cat_description_var.set(f"Description: {txn['description']}")
        self.cat_amount_var.set(f"Amount: ${abs(txn['amount']):.2f}")
        
        matching_count = sum(1 for t in self.all_transactions 
                           if t['description'] == txn['description'] and t['category'] == "Uncategorized")
        
        if matching_count > 1:
            self.cat_amount_var.set(f"Amount: ${abs(txn['amount']):.2f} ({matching_count} similar transactions)")
        
        for widget in self.category_button_frame.winfo_children():
            widget.destroy()
        
        if txn['amount'] >= 0:
            ttk.Label(self.category_button_frame, text="Payment Categories:", 
                     font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=4, sticky="ew", pady=(5, 2), padx=5)
            
            row = 1
            for i, cat in enumerate(self.parser.PAYMENT_CATEGORIES):
                btn = ttk.Button(self.category_button_frame, text=cat, 
                               command=lambda c=cat: self.apply_category_direct(c))
                btn.grid(row=row + i//4, column=i%4, padx=5, pady=5, sticky="ew")
            
            row += (len(self.parser.PAYMENT_CATEGORIES) + 3) // 4 + 1
            ttk.Label(self.category_button_frame, text="Income Categories:", 
                     font=("Arial", 12, "bold")).grid(row=row, column=0, columnspan=4, sticky="ew", pady=(10, 2), padx=5)
            
            row += 1
            for i, cat in enumerate(self.parser.INCOME_CATEGORIES):
                btn = ttk.Button(self.category_button_frame, text=cat, 
                               command=lambda c=cat: self.apply_category_direct(c))
                btn.grid(row=row + i//4, column=i%4, padx=5, pady=5, sticky="ew")
            
            row += (len(self.parser.INCOME_CATEGORIES) + 3) // 4 + 1
            ttk.Label(self.category_button_frame, text="Other:", 
                     font=("Arial", 12, "bold")).grid(row=row, column=0, columnspan=4, sticky="ew", pady=(10, 2), padx=5)
            
            row += 1
            btn = ttk.Button(self.category_button_frame, text=self.parser.IGNORE_CATEGORY, 
                           command=lambda c=self.parser.IGNORE_CATEGORY: self.apply_category_direct(c))
            btn.grid(row=row, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
            
        else:
            ttk.Label(self.category_button_frame, text="Expense Categories:", 
                     font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=4, sticky="ew", pady=(5, 2), padx=5)
            
            row = 1
            for i, cat in enumerate(self.parser.EXPENSE_CATEGORIES):
                btn = ttk.Button(self.category_button_frame, text=cat, 
                               command=lambda c=cat: self.apply_category_direct(c))
                btn.grid(row=row + i//4, column=i%4, padx=5, pady=5, sticky="ew")
            
            row += (len(self.parser.EXPENSE_CATEGORIES) + 3) // 4 + 1
            ttk.Label(self.category_button_frame, text="Other:", 
                     font=("Arial", 12, "bold")).grid(row=row, column=0, columnspan=4, sticky="ew", pady=(10, 2), padx=5)
            
            row += 1
            btn = ttk.Button(self.category_button_frame, text=self.parser.IGNORE_CATEGORY, 
                           command=lambda c=self.parser.IGNORE_CATEGORY: self.apply_category_direct(c))
            btn.grid(row=row, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
        
        self.category_button_frame.columnconfigure(0, weight=1)
        self.category_button_frame.columnconfigure(1, weight=1)
        self.category_button_frame.columnconfigure(2, weight=1)
        self.category_button_frame.columnconfigure(3, weight=1)
        
        self.cat_progress_var.set(f"Progress: {self.current_categorization_index + 1} / {len(self.pending_categorizations)}")
    
    def apply_category_direct(self, category):
        if self.current_categorization_index >= len(self.pending_categorizations):
            return
        
        txn = self.pending_categorizations[self.current_categorization_index]
        txn['category'] = category
        
        normalized = self.parser._normalize_description(txn['description'])
        self.parser.mappings[normalized] = category
        self.parser._save_mappings()
        
        description_to_match = txn['description']
        count = 0
        for t in self.all_transactions:
            if t['description'] == description_to_match and t['category'] == "Uncategorized":
                t['category'] = category
                count += 1
        
        if count > 1:
            self.log_message(f"Applied '{category}' to {count} transactions with description: {description_to_match[:50]}...")
        
        self.current_categorization_index += 1
        self.show_current_categorization()
        self.refresh_transaction_list()
    
    def skip_categorization(self):
        if self.current_categorization_index >= len(self.pending_categorizations):
            return
        
        self.current_categorization_index += 1
        self.show_current_categorization()
        self.refresh_transaction_list()
    
    def export_summary(self):
        if not self.all_transactions:
            messagebox.showerror("Error", "No transactions to export. Please add CSV files first.")
            return
        
        uncategorized = [t for t in self.all_transactions if t['category'] == "Uncategorized"]
        if uncategorized:
            response = messagebox.askyesno("Uncategorized Transactions",
                                          f"{len(uncategorized)} transactions are uncategorized.\n"
                                          "They will be excluded from the summary.\n\n"
                                          "Continue with export?")
            if not response:
                return
        
        self.log_message("\n" + "="*50)
        self.log_message("Generating Summary...")
        self.log_message("="*50)
        
        try:
            expenses, income, payments, monthly_summary, monthly_expense_breakdown = \
                self.parser.generate_summary(self.all_transactions)
            
            output_file = self.output_file_var.get()
            self.parser.export_to_excel(expenses, income, payments, monthly_summary, 
                                       monthly_expense_breakdown, output_file)
            
            self.log_message("\n" + "="*50)
            self.log_message("SUMMARY")
            self.log_message("="*50)
            self.log_message(f"Total transactions: {len(self.all_transactions)}")
            
            ignored_count = sum(1 for t in self.all_transactions if t['category'] == self.parser.IGNORE_CATEGORY)
            if ignored_count > 0:
                self.log_message(f"Ignored transactions: {ignored_count}")
            
            if uncategorized:
                self.log_message(f"Uncategorized transactions: {len(uncategorized)}")
            
            self.log_message(f"Expenses: {len(expenses)}")
            self.log_message(f"Income: {len(income)}")
            self.log_message(f"Payments: {len(payments)}")
            self.log_message("="*50)
            
            messagebox.showinfo("Success", f"Excel file created successfully!\n\n{output_file}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
            self.log_message(f"ERROR: {str(e)}")
    
    def load_config(self):
        """Load configuration from file"""
        config_file = self.config_file_var.get()
        if not os.path.exists(config_file):
            messagebox.showerror("Error", f"File not found: {config_file}")
            return
        
        try:
            self.parser.mapping_file = config_file
            self.parser._load_config()
            
            self.refresh_category_list()
            self.refresh_mapping_list()
            
            messagebox.showinfo("Success", f"Loaded configuration from {config_file}")
            self.log_message(f"Loaded config from {config_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load config: {str(e)}")
    
    def save_config(self):
        """Save configuration to file"""
        config_file = self.config_file_var.get()
        
        try:
            self.parser.mapping_file = config_file
            self.parser.save_config()
            
            messagebox.showinfo("Success", f"Saved configuration to {config_file}")
            self.log_message(f"Saved config to {config_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config: {str(e)}")
    
    def refresh_category_list(self):
        """Refresh the category listbox"""
        if not hasattr(self, 'category_listbox'):
            return
        
        cat_type = self.cat_type_var.get()
        self.category_listbox.delete(0, tk.END)
        
        if cat_type == "Expense":
            categories = self.parser.EXPENSE_CATEGORIES
        elif cat_type == "Income":
            categories = self.parser.INCOME_CATEGORIES
        else:  # Payment
            categories = self.parser.PAYMENT_CATEGORIES
        
        for cat in sorted(categories):
            self.category_listbox.insert(tk.END, cat)
    
    def add_category(self):
        """Add a new category"""
        new_cat = self.new_category_var.get().strip()
        if not new_cat:
            messagebox.showwarning("Empty Category", "Please enter a category name")
            return
        
        cat_type = self.cat_type_var.get()
        
        if cat_type == "Expense":
            if new_cat in self.parser.EXPENSE_CATEGORIES:
                messagebox.showwarning("Duplicate", "Category already exists")
                return
            self.parser.EXPENSE_CATEGORIES.append(new_cat)
        elif cat_type == "Income":
            if new_cat in self.parser.INCOME_CATEGORIES:
                messagebox.showwarning("Duplicate", "Category already exists")
                return
            self.parser.INCOME_CATEGORIES.append(new_cat)
        else:  # Payment
            if new_cat in self.parser.PAYMENT_CATEGORIES:
                messagebox.showwarning("Duplicate", "Category already exists")
                return
            self.parser.PAYMENT_CATEGORIES.append(new_cat)
        
        self.new_category_var.set("")
        self.refresh_category_list()
        self.log_message(f"Added {cat_type} category: {new_cat}")
    
    def delete_category(self):
        """Delete selected category"""
        selection = self.category_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a category to delete")
            return
        
        cat_name = self.category_listbox.get(selection[0])
        cat_type = self.cat_type_var.get()
        
        response = messagebox.askyesno("Confirm Delete", 
                                       f"Delete category '{cat_name}'?\n\n"
                                       "Note: Existing transactions with this category will keep it,\n"
                                       "but it won't be available for new transactions.")
        if not response:
            return
        
        if cat_type == "Expense":
            self.parser.EXPENSE_CATEGORIES.remove(cat_name)
        elif cat_type == "Income":
            self.parser.INCOME_CATEGORIES.remove(cat_name)
        else:  # Payment
            self.parser.PAYMENT_CATEGORIES.remove(cat_name)
        
        self.refresh_category_list()
        self.log_message(f"Deleted {cat_type} category: {cat_name}")
    
    def refresh_mapping_list(self):
        """Refresh the mappings tree"""
        if not hasattr(self, 'mapping_tree'):
            return
        
        for item in self.mapping_tree.get_children():
            self.mapping_tree.delete(item)
        
        search_term = self.mapping_search_var.get().lower()
        count = 0
        
        for desc, cat in sorted(self.parser.mappings.items()):
            if not search_term or search_term in desc.lower() or search_term in cat.lower():
                self.mapping_tree.insert("", tk.END, values=(desc, cat))
                count += 1
        
        total = len(self.parser.mappings)
        if search_term:
            self.mapping_count_var.set(f"Showing {count} of {total} mappings")
        else:
            self.mapping_count_var.set(f"Total mappings: {total}")
    
    def delete_mapping(self):
        """Delete selected mapping"""
        selection = self.mapping_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a mapping to delete")
            return
        
        item = self.mapping_tree.item(selection[0])
        desc, cat = item['values']
        
        response = messagebox.askyesno("Confirm Delete", 
                                       f"Delete mapping?\n\nDescription: {desc}\nCategory: {cat}")
        if not response:
            return
        
        del self.parser.mappings[desc]
        self.refresh_mapping_list()
        self.log_message(f"Deleted mapping: {desc} → {cat}")
    
    def clear_all_mappings(self):
        """Clear all mappings"""
        response = messagebox.askyesno("Confirm Clear All", 
                                       f"Delete ALL {len(self.parser.mappings)} mappings?\n\n"
                                       "This cannot be undone!")
        if not response:
            return
        
        count = len(self.parser.mappings)
        self.parser.mappings.clear()
        self.refresh_mapping_list()
        self.log_message(f"Cleared {count} mappings")
        messagebox.showinfo("Cleared", f"Deleted {count} mappings")


